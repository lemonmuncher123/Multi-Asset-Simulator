import json
import tempfile
from pathlib import Path

import pytest

from src.storage.database import init_db
from src.models.asset import Asset
from src.models.transaction import Transaction
from src.storage.asset_repo import create_asset
from src.storage.transaction_repo import create_transaction
from src.engines.reports import build_period_report, generate_monthly_report
from src.engines.report_export import export_report_txt, export_report_xlsx


@pytest.fixture
def db_conn():
    conn = init_db(":memory:")
    yield conn
    conn.close()


@pytest.fixture
def populated_db(db_conn):
    a = create_asset(db_conn, Asset(symbol="AAPL", name="Apple", asset_type="stock"))
    create_transaction(db_conn, Transaction(
        date="2025-06-01", txn_type="deposit_cash",
        total_amount=100000.0, currency="USD",
    ))
    create_transaction(db_conn, Transaction(
        date="2025-06-15", txn_type="buy", asset_id=a.id,
        quantity=10, price=150.0, total_amount=-1500.0, currency="USD", fees=10.0,
    ))
    return db_conn


@pytest.fixture
def sample_report_data(populated_db):
    report = generate_monthly_report(populated_db, 2025, 6)
    return json.loads(report.report_json)


@pytest.fixture
def empty_report_data(db_conn):
    report = build_period_report(db_conn, "2099-01-01", "2099-02-01", "2099-01", "monthly")
    return json.loads(report.report_json)


# ===================================================================
# TXT export
# ===================================================================

class TestTxtExport:

    def test_creates_readable_file(self, sample_report_data):
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "report.txt"
            export_report_txt(sample_report_data, path)
            assert path.exists()
            content = path.read_text()
            assert len(content) > 100

    def test_contains_how_to_read(self, sample_report_data):
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "report.txt"
            export_report_txt(sample_report_data, path)
            content = path.read_text()
            assert "HOW TO READ THIS REPORT" in content

    def test_contains_net_cash_flow(self, sample_report_data):
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "report.txt"
            export_report_txt(sample_report_data, path)
            content = path.read_text()
            assert "Net Cash Flow" in content

    def test_contains_operating_net_income(self, sample_report_data):
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "report.txt"
            export_report_txt(sample_report_data, path)
            content = path.read_text()
            assert "Operating Net Income" in content

    def test_contains_period_label(self, sample_report_data):
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "report.txt"
            export_report_txt(sample_report_data, path)
            content = path.read_text()
            assert "MONTHLY REPORT: 2025-06" in content

    def test_contains_summary_section(self, sample_report_data):
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "report.txt"
            export_report_txt(sample_report_data, path)
            content = path.read_text()
            assert "SUMMARY" in content
            assert "Beginning Cash" in content
            assert "Ending Cash" in content

    def test_contains_snapshot_section(self, sample_report_data):
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "report.txt"
            export_report_txt(sample_report_data, path)
            content = path.read_text()
            assert "CURRENT SNAPSHOT" in content

    def test_empty_report_still_exports(self, empty_report_data):
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "empty_report.txt"
            export_report_txt(empty_report_data, path)
            assert path.exists()
            content = path.read_text()
            assert "HOW TO READ THIS REPORT" in content
            assert "Transaction Count:   0" in content


# ===================================================================
# XLSX export
# ===================================================================

class TestXlsxExport:

    def test_creates_workbook(self, sample_report_data):
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "report.xlsx"
            export_report_xlsx(sample_report_data, path)
            assert path.exists()
            assert path.stat().st_size > 0

    def test_contains_all_required_sheets(self, sample_report_data):
        import openpyxl
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "report.xlsx"
            export_report_xlsx(sample_report_data, path)
            wb = openpyxl.load_workbook(path)
            sheet_names = wb.sheetnames
            required = [
                "How To Read",
                "Summary",
                "Operations",
                "Transactions",
                "Trades",
                "Real Estate",
                "Debt",
                "Journal",
                "Current Snapshot",
            ]
            for name in required:
                assert name in sheet_names, f"Missing sheet: {name}"
            wb.close()

    def test_summary_values_match_report_json(self, sample_report_data):
        import openpyxl
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "report.xlsx"
            export_report_xlsx(sample_report_data, path)
            wb = openpyxl.load_workbook(path)
            ws = wb["Summary"]
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            metrics = {row[0]: row[1] for row in rows}

            s = sample_report_data["summary"]
            assert metrics["Report Type"] == s["report_type"]
            assert metrics["Period Label"] == s["period_label"]
            assert metrics["Transaction Count"] == s["transaction_count"]
            assert metrics["Net Cash Flow"] == s["net_cash_flow"]
            assert metrics["Operating Net Income"] == s["operating_net_income"]
            assert metrics["Total Fees"] == s["total_fees"]
            wb.close()

    def test_transactions_sheet_has_data(self, sample_report_data):
        import openpyxl
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "report.xlsx"
            export_report_xlsx(sample_report_data, path)
            wb = openpyxl.load_workbook(path)
            ws = wb["Transactions"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            assert len(data_rows) == 2
            wb.close()

    def test_empty_report_xlsx(self, empty_report_data):
        import openpyxl
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "empty.xlsx"
            export_report_xlsx(empty_report_data, path)
            assert path.exists()
            wb = openpyxl.load_workbook(path)
            assert "Summary" in wb.sheetnames
            ws = wb["Summary"]
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            metrics = {row[0]: row[1] for row in rows}
            assert metrics["Transaction Count"] == 0
            wb.close()

    def test_current_snapshot_sheet(self, sample_report_data):
        import openpyxl
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "report.xlsx"
            export_report_xlsx(sample_report_data, path)
            wb = openpyxl.load_workbook(path)
            ws = wb["Current Snapshot"]
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            metrics = {row[0]: row[1] for row in rows}
            assert "Cash" in metrics
            assert "Net Worth" in metrics
            wb.close()

    def test_how_to_read_sheet_not_empty(self, sample_report_data):
        import openpyxl
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "report.xlsx"
            export_report_xlsx(sample_report_data, path)
            wb = openpyxl.load_workbook(path)
            ws = wb["How To Read"]
            rows = list(ws.iter_rows(values_only=True))
            assert len(rows) > 5
            wb.close()


# ===================================================================
# Phase 1: TXT export — updated explanation, new sections
# ===================================================================


class TestTxtExportPhase1:

    def test_txt_explains_cash_movement_vs_profit(self, sample_report_data):
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "report.txt"
            export_report_txt(sample_report_data, path)
            content = path.read_text().lower()
            assert "cash movement" in content
            assert "not profit" in content

    def test_txt_explains_funding_flow(self, sample_report_data):
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "report.txt"
            export_report_txt(sample_report_data, path)
            content = path.read_text().lower()
            assert "funding flow" in content
            # Funding is described as funding (not income).
            assert "not income" in content or "funding, not income" in content

    def test_txt_explains_approximate_investment_result(self, sample_report_data):
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "report.txt"
            export_report_txt(sample_report_data, path)
            content = path.read_text()
            assert "Approximate Investment Result" in content
            assert "Net Worth Change" in content

    def test_txt_explains_snapshot_fallback(self, sample_report_data):
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "report.txt"
            export_report_txt(sample_report_data, path)
            content = path.read_text().lower()
            assert "snapshot" in content
            # Mentions either "at or before" semantics or fallback wording.
            assert "at or before" in content or "no snapshot" in content

    def test_txt_contains_cash_flow_breakdown_section(self, sample_report_data):
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "report.txt"
            export_report_txt(sample_report_data, path)
            content = path.read_text()
            assert "CASH FLOW BREAKDOWN" in content
            assert "Funding Flow" in content
            assert "Trade Cash Flow" in content
            assert "Real Estate Cash Flow" in content
            assert "Debt Cash Flow" in content

    def test_txt_contains_performance_section(self, sample_report_data):
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "report.txt"
            export_report_txt(sample_report_data, path)
            content = path.read_text()
            assert "PERFORMANCE" in content
            assert "Beginning Net Worth" in content
            assert "Ending Net Worth" in content
            assert "Approximate Investment Result" in content
            assert "Approximate Return %" in content

    def test_txt_does_not_label_as_strict_twr_or_irr(self, sample_report_data):
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "report.txt"
            export_report_txt(sample_report_data, path)
            content = path.read_text().lower()
            # No claim of strict TWR / IRR / time-weighted-return calculation.
            assert "time-weighted return" not in content
            assert "internal rate of return" not in content

    def test_txt_handles_old_report_without_new_sections(self):
        old_report = {
            "summary": {
                "report_type": "monthly",
                "period_label": "2024-01",
                "period_start": "2024-01-01",
                "period_end": "2024-02-01",
                "generated_at": "2024-02-01T00:00:00",
                "transaction_count": 0,
                "beginning_cash": 0.0,
                "ending_cash": 0.0,
                "net_cash_flow": 0.0,
                "operating_net_income": 0.0,
                "total_inflow": 0.0,
                "total_outflow": 0.0,
                "total_fees": 0.0,
            },
            "operations": [],
            "transactions": [],
            "trades": [],
            "real_estate": [],
            "debt": [],
            "journal": [],
            "current_snapshot": {
                "note": "x",
                "cash": 0.0,
                "total_assets": 0.0,
                "total_liabilities": 0.0,
                "net_worth": 0.0,
            },
        }
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "old.txt"
            export_report_txt(old_report, path)
            assert path.exists()
            content = path.read_text()
            # HOW TO READ is still emitted; missing sections are skipped, not crashed.
            assert "HOW TO READ THIS REPORT" in content


# ===================================================================
# Phase 1: XLSX export — new sheets
# ===================================================================


class TestXlsxExportPhase1:

    def test_xlsx_has_cash_flow_breakdown_sheet(self, sample_report_data):
        import openpyxl
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "report.xlsx"
            export_report_xlsx(sample_report_data, path)
            wb = openpyxl.load_workbook(path)
            assert "Cash Flow Breakdown" in wb.sheetnames
            wb.close()

    def test_xlsx_has_performance_sheet(self, sample_report_data):
        import openpyxl
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "report.xlsx"
            export_report_xlsx(sample_report_data, path)
            wb = openpyxl.load_workbook(path)
            assert "Performance" in wb.sheetnames
            wb.close()

    def test_xlsx_cfb_sheet_has_required_categories(self, sample_report_data):
        import openpyxl
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "report.xlsx"
            export_report_xlsx(sample_report_data, path)
            wb = openpyxl.load_workbook(path)
            ws = wb["Cash Flow Breakdown"]
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            categories = {row[0] for row in rows}
            for required in (
                "Funding Flow", "Trade Cash Flow", "Real Estate Cash Flow",
                "Debt Cash Flow", "Fees Total", "Other Cash Flow",
            ):
                assert required in categories
            wb.close()

    def test_xlsx_performance_sheet_has_expected_metrics(self, sample_report_data):
        import openpyxl
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "report.xlsx"
            export_report_xlsx(sample_report_data, path)
            wb = openpyxl.load_workbook(path)
            ws = wb["Performance"]
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            metrics = {row[0]: row[1] for row in rows}
            for key in (
                "Beginning Net Worth", "Ending Net Worth", "Net Worth Change",
                "Funding Flow", "Approximate Investment Result",
                "Approximate Return %", "Data Quality Note",
            ):
                assert key in metrics
            wb.close()

    def test_xlsx_empty_report_still_has_new_sheets(self, empty_report_data):
        import openpyxl
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "empty.xlsx"
            export_report_xlsx(empty_report_data, path)
            wb = openpyxl.load_workbook(path)
            assert "Cash Flow Breakdown" in wb.sheetnames
            assert "Performance" in wb.sheetnames
            wb.close()

    def test_xlsx_handles_old_report_without_new_sections(self):
        import openpyxl
        old_report = {
            "summary": {
                "report_type": "monthly",
                "period_label": "2024-01",
                "period_start": "2024-01-01",
                "period_end": "2024-02-01",
                "generated_at": "2024-02-01T00:00:00",
                "transaction_count": 0,
                "beginning_cash": 0.0,
                "ending_cash": 0.0,
                "net_cash_flow": 0.0,
                "operating_net_income": 0.0,
                "total_inflow": 0.0,
                "total_outflow": 0.0,
                "total_fees": 0.0,
            },
            "operations": [],
            "transactions": [],
            "trades": [],
            "real_estate": [],
            "debt": [],
            "journal": [],
            "current_snapshot": {
                "note": "x",
                "cash": 0.0,
                "total_assets": 0.0,
                "total_liabilities": 0.0,
                "net_worth": 0.0,
            },
        }
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "old.xlsx"
            export_report_xlsx(old_report, path)
            wb = openpyxl.load_workbook(path)
            # New sheets are emitted with empty rows / N/A values, not crashed.
            assert "Cash Flow Breakdown" in wb.sheetnames
            assert "Performance" in wb.sheetnames
            wb.close()


# ===================================================================
# Phase 2: TXT export — Allocation + Risk Summary sections
# ===================================================================


class TestPhase2TxtExport:

    def test_txt_contains_allocation_section_header(self, sample_report_data):
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "report.txt"
            export_report_txt(sample_report_data, path)
            content = path.read_text()
            assert "ALLOCATION" in content

    def test_txt_allocation_includes_source_and_balance_lines(self, sample_report_data):
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "report.txt"
            export_report_txt(sample_report_data, path)
            content = path.read_text()
            assert "Source:" in content
            assert "Cash Amount" in content
            assert "Total Assets" in content
            assert "Net Worth" in content

    def test_txt_allocation_includes_data_quality_note(self, sample_report_data):
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "report.txt"
            export_report_txt(sample_report_data, path)
            content = path.read_text()
            note = sample_report_data["allocation"]["data_quality_note"]
            assert note
            assert note in content

    def test_txt_contains_risk_summary_header_and_total(self, sample_report_data):
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "report.txt"
            export_report_txt(sample_report_data, path)
            content = path.read_text()
            assert "RISK SUMMARY" in content
            assert "Total:" in content

    def test_txt_risk_marks_observations_not_recommendations(self, sample_report_data):
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "report.txt"
            export_report_txt(sample_report_data, path)
            content = path.read_text().lower()
            # Risk warnings should use observational language, not financial advice language.
            assert "observations" in content
            assert "not recommendations" in content

    def test_txt_how_to_read_explains_allocation_and_risk(self, sample_report_data):
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "report.txt"
            export_report_txt(sample_report_data, path)
            content = path.read_text()
            assert "ALLOCATION:" in content
            assert "RISK SUMMARY:" in content

    def test_txt_handles_old_report_without_alloc_or_risk(self):
        old_report = {
            "summary": {
                "report_type": "monthly",
                "period_label": "2024-01",
                "period_start": "2024-01-01",
                "period_end": "2024-02-01",
                "generated_at": "2024-02-01T00:00:00",
                "transaction_count": 0,
                "beginning_cash": 0.0,
                "ending_cash": 0.0,
                "net_cash_flow": 0.0,
                "operating_net_income": 0.0,
                "total_inflow": 0.0,
                "total_outflow": 0.0,
                "total_fees": 0.0,
            },
            "operations": [],
            "transactions": [],
            "trades": [],
            "real_estate": [],
            "debt": [],
            "journal": [],
            "current_snapshot": {
                "note": "x",
                "cash": 0.0,
                "total_assets": 0.0,
                "total_liabilities": 0.0,
                "net_worth": 0.0,
            },
        }
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "old.txt"
            # Must not crash; missing sections are simply skipped.
            export_report_txt(old_report, path)
            assert path.exists()


# ===================================================================
# Phase 2: XLSX export — Allocation + Risk Summary sheets
# ===================================================================


class TestPhase2XlsxExport:

    def test_xlsx_has_allocation_sheet(self, sample_report_data):
        import openpyxl
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "report.xlsx"
            export_report_xlsx(sample_report_data, path)
            wb = openpyxl.load_workbook(path)
            assert "Allocation" in wb.sheetnames
            wb.close()

    def test_xlsx_has_risk_summary_sheet(self, sample_report_data):
        import openpyxl
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "report.xlsx"
            export_report_xlsx(sample_report_data, path)
            wb = openpyxl.load_workbook(path)
            assert "Risk Summary" in wb.sheetnames
            wb.close()

    def test_xlsx_allocation_sheet_has_required_metrics(self, sample_report_data):
        import openpyxl
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "report.xlsx"
            export_report_xlsx(sample_report_data, path)
            wb = openpyxl.load_workbook(path)
            ws = wb["Allocation"]
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            metrics = {row[0]: row[1] for row in rows}
            for required in (
                "Source", "Cash Amount", "Total Assets",
                "Total Liabilities", "Net Worth", "Data Quality Note",
            ):
                assert required in metrics
            wb.close()

    def test_xlsx_risk_summary_sheet_has_counts_and_note(self, sample_report_data):
        import openpyxl
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "report.xlsx"
            export_report_xlsx(sample_report_data, path)
            wb = openpyxl.load_workbook(path)
            ws = wb["Risk Summary"]
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            metrics = {row[0]: row[1] for row in rows}
            for required in (
                "Total Count", "Warning Count (actionable)", "Info Count",
                "Data Quality Note",
            ):
                assert required in metrics
            wb.close()

    def test_xlsx_empty_report_has_new_sheets(self, empty_report_data):
        import openpyxl
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "empty.xlsx"
            export_report_xlsx(empty_report_data, path)
            wb = openpyxl.load_workbook(path)
            assert "Allocation" in wb.sheetnames
            assert "Risk Summary" in wb.sheetnames
            wb.close()

    def test_xlsx_handles_old_report_without_alloc_or_risk(self):
        import openpyxl
        old_report = {
            "summary": {
                "report_type": "monthly",
                "period_label": "2024-01",
                "period_start": "2024-01-01",
                "period_end": "2024-02-01",
                "generated_at": "2024-02-01T00:00:00",
                "transaction_count": 0,
                "beginning_cash": 0.0,
                "ending_cash": 0.0,
                "net_cash_flow": 0.0,
                "operating_net_income": 0.0,
                "total_inflow": 0.0,
                "total_outflow": 0.0,
                "total_fees": 0.0,
            },
            "operations": [],
            "transactions": [],
            "trades": [],
            "real_estate": [],
            "debt": [],
            "journal": [],
            "current_snapshot": {
                "note": "x",
                "cash": 0.0,
                "total_assets": 0.0,
                "total_liabilities": 0.0,
                "net_worth": 0.0,
            },
        }
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "old.xlsx"
            export_report_xlsx(old_report, path)
            wb = openpyxl.load_workbook(path)
            # Sheets are emitted with default values, not crashed.
            assert "Allocation" in wb.sheetnames
            assert "Risk Summary" in wb.sheetnames
            wb.close()


# ===================================================================
# Beginning/Ending Snapshot + Fees Breakdown rendering
# ===================================================================


@pytest.fixture
def sample_report_data_with_fees(populated_db):
    """Same as populated_db but with a transaction_fee_breakdown row to
    populate the fees_breakdown section."""
    from src.storage.fee_breakdown_repo import (
        create_fee_breakdown, FeeBreakdownRow,
    )
    txn_id = populated_db.execute(
        "SELECT id FROM transactions WHERE txn_type='buy' ORDER BY id DESC LIMIT 1"
    ).fetchone()[0]
    create_fee_breakdown(populated_db, FeeBreakdownRow(
        transaction_id=txn_id, fee_type="broker_commission", amount=10.0,
    ))
    report = generate_monthly_report(populated_db, 2025, 6)
    return json.loads(report.report_json)


class TestTxtExportBeginningEndingFees:

    def test_txt_includes_beginning_and_ending_snapshot_headers(self, sample_report_data):
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "r.txt"
            export_report_txt(sample_report_data, path)
            text = path.read_text()
            assert "BEGINNING SNAPSHOT" in text
            assert "ENDING SNAPSHOT" in text

    def test_txt_includes_fees_breakdown_when_present(self, sample_report_data_with_fees):
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "r.txt"
            export_report_txt(sample_report_data_with_fees, path)
            text = path.read_text()
            assert "FEES BREAKDOWN" in text
            assert "broker_commission" in text

    def test_txt_omits_fees_breakdown_when_empty(self, sample_report_data):
        """The FEES BREAKDOWN section (its rendered table) is omitted when
        empty. The 'FEES BREAKDOWN:' bullet in HOW_TO_READ stays — that's
        documentation, not the section."""
        sample_report_data["fees_breakdown"] = {"by_type": [], "grand_total": 0.0}
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "r.txt"
            export_report_txt(sample_report_data, path)
            text = path.read_text()
            # The rendered section header is exactly "FEES BREAKDOWN" on its
            # own line followed by dashes. The HOW_TO_READ bullet is
            # "FEES BREAKDOWN:" with a colon. Discriminate on the trailing
            # dash separator.
            assert "FEES BREAKDOWN\n" + ("-" * 40) not in text
            # And no 'GRAND TOTAL' line when empty.
            assert "GRAND TOTAL" not in text

    def test_txt_how_to_read_mentions_new_sections(self, sample_report_data):
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "r.txt"
            export_report_txt(sample_report_data, path)
            text = path.read_text()
            assert "SNAPSHOTS (BEGINNING / ENDING)" in text
            assert "FEES BREAKDOWN:" in text


class TestXlsxExportBeginningEndingFees:

    def test_xlsx_has_snapshots_sheet(self, sample_report_data):
        import openpyxl
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "r.xlsx"
            export_report_xlsx(sample_report_data, path)
            wb = openpyxl.load_workbook(path)
            assert "Snapshots" in wb.sheetnames
            wb.close()

    def test_xlsx_snapshots_sheet_has_beg_and_end_rows(self, sample_report_data):
        import openpyxl
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "r.xlsx"
            export_report_xlsx(sample_report_data, path)
            wb = openpyxl.load_workbook(path)
            ws = wb["Snapshots"]
            metrics = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
            assert any("Beginning" in (m or "") for m in metrics)
            assert any("Ending" in (m or "") for m in metrics)
            wb.close()

    def test_xlsx_has_fees_breakdown_sheet(self, sample_report_data_with_fees):
        import openpyxl
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "r.xlsx"
            export_report_xlsx(sample_report_data_with_fees, path)
            wb = openpyxl.load_workbook(path)
            assert "Fees Breakdown" in wb.sheetnames
            ws = wb["Fees Breakdown"]
            assert ws.max_row >= 2  # header + at least one fee row
            wb.close()

    def test_xlsx_fees_breakdown_sheet_present_even_when_empty(self, sample_report_data):
        """Discoverability: empty Fees Breakdown sheet is still emitted."""
        import openpyxl
        sample_report_data["fees_breakdown"] = {"by_type": [], "grand_total": 0.0}
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "r.xlsx"
            export_report_xlsx(sample_report_data, path)
            wb = openpyxl.load_workbook(path)
            assert "Fees Breakdown" in wb.sheetnames
            wb.close()
